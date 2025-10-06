import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl.styles import NamedStyle
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook

st.title("Extract Rows by Names from Excel (All Sheets) with ISIN Autofill (Column Index Lookup)")

uploaded_file = st.file_uploader("Upload Excel file (any number of sheets)", type=["xlsx"])
lookup_file = st.file_uploader("Upload ISIN Lookup Excel file (no headers, fixed columns)", type=["xlsx"])

if uploaded_file:
    xls = pd.ExcelFile(uploaded_file)
    sheet_names = xls.sheet_names
    sheets_to_use = sheet_names   # ✅ Now all sheets are considered

    # Collect unique names from all sheets
    all_names = set()
    for sheet in sheets_to_use:
        df = pd.read_excel(xls, sheet_name=sheet, usecols="A:J")
        all_names.update(df['Name'].dropna().astype(str).unique())
    all_names = sorted(all_names)

    selected_names = st.multiselect("Select one or more names:", all_names)

    lookup_df = None
    if lookup_file:
        # Read lookup file with no headers, usecols by column numbers
        lookup_df = pd.read_excel(
            lookup_file,
            header=None,
            usecols=[0, 1, 5, 7]  # A, B, F, H
        )
        lookup_df.columns = ['nbfc', 'isin', 'issue_date', 'maturity_date']

        # Convert dates to datetime
        lookup_df['issue_date'] = pd.to_datetime(lookup_df['issue_date'], errors='coerce')
        lookup_df['maturity_date'] = pd.to_datetime(lookup_df['maturity_date'], errors='coerce')

    if selected_names:
        combined_rows = []

        for sheet in sheets_to_use:
            df = pd.read_excel(xls, sheet_name=sheet, usecols="A:J")

            filtered = df[(df['Name'].astype(str).isin(selected_names)) & 
                          (pd.to_numeric(df['Units'], errors='coerce') > 0)]

            if not filtered.empty:
                # Drop empty or unnamed columns
                cols_to_drop = [col for col in filtered.columns 
                                if (isinstance(col, str) and (col.strip() == '' or col.startswith('Unnamed')))]
                filtered = filtered.drop(columns=cols_to_drop)

                # Insert empty ISIN column after 'NBFC'
                if 'NBFC' in filtered.columns:
                    idx = filtered.columns.get_loc('NBFC') + 1
                    filtered.insert(idx, 'ISIN', '')

                if lookup_df is not None:
                    # Convert filtered dates to datetime
                    if 'Issue Date' in filtered.columns and 'Maturity Date' in filtered.columns:
                        filtered['Issue Date'] = pd.to_datetime(filtered['Issue Date'], errors='coerce')
                        filtered['Maturity Date'] = pd.to_datetime(filtered['Maturity Date'], errors='coerce')

                        def get_isin(row):
                            match = lookup_df[
                                (lookup_df['nbfc'] == row['NBFC']) &
                                (lookup_df['issue_date'] == row['Issue Date']) &
                                (lookup_df['maturity_date'] == row['Maturity Date'])
                            ]
                            return match.iloc[0]['isin'] if not match.empty else ''

                        filtered['ISIN'] = filtered.apply(get_isin, axis=1)

                combined_rows.append(filtered)

        if combined_rows:
            final_df = pd.concat(combined_rows, ignore_index=True)

            # Create Excel file with formatting
            output = BytesIO()
            wb = Workbook()
            ws = wb.active
            ws.title = 'Extracted'

            # Add headers and rows
            for r in dataframe_to_rows(final_df, index=False, header=True):
                ws.append(r)

            # Create styles
            percentage_style = NamedStyle(name="percentage")
            percentage_style.number_format = '0.00%'

            comma_style = NamedStyle(name="comma")
            comma_style.number_format = '#,##0'

            date_style = NamedStyle(name="date_dd_mmm_yy")
            date_style.number_format = 'DD-MMM-YY'

            month_year_style = NamedStyle(name="month_year")
            month_year_style.number_format = 'MMM-YY'

            # Register styles in workbook
            for style in [percentage_style, comma_style, date_style, month_year_style]:
                if style.name not in wb.named_styles:
                    wb.add_named_style(style)

            # Apply formatting based on column names
            for col_idx, col_name in enumerate(final_df.columns, 1):
                col_letter = ws.cell(row=1, column=col_idx).column_letter

                if 'Interest' in str(col_name):
                    for row in range(2, ws.max_row + 1):
                        cell = ws[f'{col_letter}{row}']
                        if cell.value is not None:
                            try:
                                if isinstance(cell.value, (int, float)) and cell.value > 1:
                                    cell.value = cell.value / 100
                                cell.style = percentage_style
                            except:
                                pass

                elif 'Amount' in str(col_name):
                    for row in range(2, ws.max_row + 1):
                        cell = ws[f'{col_letter}{row}']
                        if cell.value is not None:
                            try:
                                cell.style = comma_style
                            except:
                                pass

                elif 'Date' in str(col_name) and not any(x in str(col_name).lower() for x in ['issue', 'maturity']):
                    for row in range(2, ws.max_row + 1):
                        cell = ws[f'{col_letter}{row}']
                        if cell.value is not None:
                            try:
                                cell.style = date_style
                            except:
                                pass

                elif any(x in str(col_name).lower() for x in ['issue date', 'maturity date', 'issue', 'maturity']):
                    for row in range(2, ws.max_row + 1):
                        cell = ws[f'{col_letter}{row}']
                        if cell.value is not None:
                            try:
                                cell.style = month_year_style
                            except:
                                pass

            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            wb.save(output)
            output.seek(0)

            st.download_button(
                label="Download Extracted Excel",
                data=output.getvalue(),
                file_name="extracted_rows_with_isin.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        else:
            st.info("No matching rows found with positive Units.")
